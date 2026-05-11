#!/usr/bin/env python3
"""
rename_sb_to_pubid.py

Renames SB County document folders and all files within them from USGS Index IDs
to numeric Publication IDs, matching the naming convention of all other docs.

Also:
  - Rewrites per-page metadata CSVs to use the Publication ID in the 'id' column
    and appends USGS doc-level fields from USGS_ID.xlsx
  - Updates the 'id' column in the main metadata CSV for all SB rows

ID mapping (Index ID -> Publication ID):
    ofr5983  -> 23894    ofr6196  -> 23879    ofr49118 -> 52056
    ofr64117 -> 23993    ofr6291  -> 23996    ofr63103 -> 23997
    ofr60102 -> 52119
"""

import csv
import shutil
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

REPO_ROOT   = Path(__file__).resolve().parents[3]
OUTPUT_ROOT = REPO_ROOT / "data" / "digitized"
MAIN_META   = REPO_ROOT / "Data_Files" / "cleaned_metadata_final - Copy.csv"
USGS_XLSX   = (REPO_ROOT / "Chapter_2_USGS_Digitization" / "Literature-Data Review"
               / "Edited USGS Data Pulls" / "usgs_to_id" / "USGS_ID.xlsx")

# Index ID -> Publication ID
ID_MAP = {
    "ofr5983":  "23894",
    "ofr6196":  "23879",
    "ofr49118": "52056",
    "ofr64117": "23993",
    "ofr6291":  "23996",
    "ofr63103": "23997",
    "ofr60102": "52119",
}


def load_usgs_crosswalk(path: Path) -> dict:
    """Return dict: index_id -> row dict from USGS_ID.xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value) if c.value is not None else ""
               for c in next(ws.iter_rows(min_row=1, max_row=1))]
    crosswalk = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {h: (str(v) if v is not None else "") for h, v in zip(headers, row)}
        index_id = row_dict.get("Index ID", "").strip()
        if index_id:
            crosswalk[index_id] = row_dict
    wb.close()
    return crosswalk


def rewrite_metadata_csv(path: Path, old_id: str, new_id: str,
                         usgs_row: dict, meta_cols: list) -> None:
    """Read metadata CSV, update id field, add USGS fields, write back."""
    with open(path, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    usgs_extra = [k for k in usgs_row if k not in meta_cols and k]
    all_fields = meta_cols + usgs_extra

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=all_fields, extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            updated = dict(r)
            updated["id"] = new_id
            writer.writerow({**updated, **usgs_row})


def main() -> None:
    print("Loading USGS crosswalk...", flush=True)
    crosswalk = load_usgs_crosswalk(USGS_XLSX)

    # Verify all SB docs are in the crosswalk
    for old_id in ID_MAP:
        if old_id not in crosswalk:
            print(f"WARNING: {old_id} not found in crosswalk")

    # Get metadata column list from the main metadata header
    with open(MAIN_META, encoding="utf-8") as f:
        meta_cols = csv.DictReader(f).fieldnames

    stats = {"folders": 0, "files": 0, "meta_csvs": 0}

    # ── 1. Rename files and rewrite metadata CSVs ─────────────────────────────
    for old_id, new_id in ID_MAP.items():
        doc_dir = OUTPUT_ROOT / old_id
        if not doc_dir.exists():
            print(f"WARNING: {doc_dir} not found, skipping")
            continue

        usgs_row = crosswalk.get(old_id, {})
        print(f"{old_id} -> {new_id} ...", flush=True)

        for page_dir in sorted(doc_dir.iterdir()):
            if not page_dir.is_dir():
                continue
            for file_path in sorted(page_dir.iterdir()):
                if not file_path.is_file():
                    continue

                # Rewrite metadata CSV content before renaming (use current name)
                if file_path.name.endswith("_metadata.csv") and old_id in file_path.name:
                    rewrite_metadata_csv(file_path, old_id, new_id, usgs_row, meta_cols)
                    stats["meta_csvs"] += 1

                # Rename file: replace old_id prefix with new_id (skip if already done)
                new_name = file_path.name.replace(f"{old_id}_", f"{new_id}_", 1)
                if new_name != file_path.name:
                    file_path.rename(page_dir / new_name)
                    stats["files"] += 1

        # ── 2. Rename folder via copy+delete (avoids Dropbox lock on rename) ──
        new_dir = OUTPUT_ROOT / new_id
        if not new_dir.exists():
            shutil.copytree(doc_dir, new_dir)
            shutil.rmtree(doc_dir)
        stats["folders"] += 1

    # ── 3. Update main metadata CSV ─────────────────────────────────────────
    print("Updating main metadata CSV...", flush=True)
    old_ids = set(ID_MAP.keys())

    with open(MAIN_META, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        all_cols = reader.fieldnames
        rows = list(reader)

    changed = 0
    for row in rows:
        if row["id"] in old_ids:
            row["id"] = ID_MAP[row["id"]]
            changed += 1

    with open(MAIN_META, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=all_cols, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    print(f"\nDone.")
    print(f"  Folders renamed:       {stats['folders']}")
    print(f"  Files renamed:         {stats['files']}")
    print(f"  Metadata CSVs updated: {stats['meta_csvs']}")
    print(f"  Master metadata rows updated: {changed}")


if __name__ == "__main__":
    main()
