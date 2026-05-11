#!/usr/bin/env python3
"""
organize_data.py

Reorganizes digitized USGS water data files into the target per-page structure:

    data/digitized/{doc_id}/page_{N}/
        {doc_id}_page_{N}.png
        {doc_id}_page_{N}.json
        {doc_id}_page_{N}_table1.csv  [+ table2.csv, etc.]
        {doc_id}_page_{N}_metadata.csv

Sources (relative to repo root):
    CSVs:      Data_Files/ReductCSVs/{doc_id}/{doc_id}_page_{N}_table{M}.csv
    JSONs:     Data_Files/ReductJson/{doc_id}/{doc_id}_page_{N}.json
    PNGs:      Data_Files/UpdatedDataJan/{doc_id}/ (UpdatedDataDec/ as fallback)
    Metadata:  Data_Files/cleaned_metadata_final - Copy.csv
    Crosswalk: Chapter_2_USGS_Digitization/.../usgs_to_id/USGS_ID.xlsx

Per-page metadata CSV contains all rows from the main metadata for that
(doc_id, page_number) pair, with USGS document-level fields (title, year,
author, URL, etc.) joined from USGS_ID.xlsx via doc_id = Publication ID.

Santa Barbara County data (updatedSB/Scanned_SB/ and Data_Files/Scanned_SB/)
uses a different naming convention and is not included in the main metadata.
It is excluded here and flagged in the log for separate handling.

Usage:
    python organize_data.py                  # process all docs
    python organize_data.py --dry-run        # preview without copying files
    python organize_data.py --doc-id 2252    # process a single doc (for testing)
"""

import argparse
import csv
import re
import shutil
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────────────────────────────
# Script lives at: code/01_digitization/06_final_data_organization/organize_data.py
REPO_ROOT   = Path(__file__).resolve().parents[3]
DATA_FILES  = REPO_ROOT / "Data_Files"
OUTPUT_ROOT = REPO_ROOT / "data" / "digitized"

CSV_DIR  = DATA_FILES / "ReductCSVs"
JSON_DIR = DATA_FILES / "ReductJson"
PNG_DIRS = [
    DATA_FILES / "UpdatedDataJan",   # check Jan first (more recent batch)
    DATA_FILES / "UpdatedDataDec",
]
METADATA_CSV = DATA_FILES / "cleaned_metadata_final - Copy.csv"
USGS_XLSX = (
    REPO_ROOT
    / "Chapter_2_USGS_Digitization"
    / "Literature-Data Review"
    / "Edited USGS Data Pulls"
    / "usgs_to_id"
    / "USGS_ID.xlsx"
)

# Doc IDs that are SB County data with non-standard naming — skip for now
SB_PREFIXES = ("sb",)


def load_metadata(path: Path) -> dict:
    """Return dict: (doc_id, page_num) -> list of row dicts from main metadata CSV."""
    meta = defaultdict(list)
    with open(path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            doc_id = str(row["id"]).strip()
            page   = str(row["page_number"]).strip()
            if doc_id:
                meta[(doc_id, page)].append(dict(row))
    return meta


def load_usgs_crosswalk(path: Path) -> dict:
    """Return dict: Publication ID string -> row dict from USGS_ID.xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [
        str(c.value) if c.value is not None else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    crosswalk = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {h: (str(v) if v is not None else "") for h, v in zip(headers, row)}
        # IDs may be stored as floats (e.g., "2252.0") — normalise to integer string
        pub_id = str(row_dict.get("Publication ID", "")).split(".")[0].strip()
        if pub_id:
            crosswalk[pub_id] = row_dict
    wb.close()
    return crosswalk


def find_png(doc_id: str, page_num: str) -> Path | None:
    """Return path to PNG file, checking UpdatedDataJan then UpdatedDataDec."""
    filename = f"{doc_id}_page_{page_num}.png"
    for png_dir in PNG_DIRS:
        candidate = png_dir / doc_id / filename
        if candidate.exists():
            return candidate
    return None


def copy_file(src: Path, dst: Path, dry_run: bool) -> None:
    if not dry_run:
        shutil.copy2(src, dst)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Organize USGS digitized data into per-page folder structure."
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview actions without copying any files."
    )
    parser.add_argument(
        "--doc-id",
        help="Process only this one doc ID (useful for testing)."
    )
    args = parser.parse_args()
    dry = args.dry_run

    if dry:
        print("DRY RUN — no files will be copied.\n")

    print("Loading main metadata...", flush=True)
    meta_lookup = load_metadata(METADATA_CSV)

    print("Loading USGS crosswalk...", flush=True)
    usgs_crosswalk = load_usgs_crosswalk(USGS_XLSX)

    if not dry:
        OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)

    # Enumerate all doc ID folders from the JSON directory (one JSON per page)
    all_doc_ids = sorted(d.name for d in JSON_DIR.iterdir() if d.is_dir())

    # Identify and skip SB County docs (non-standard naming)
    sb_doc_ids   = [d for d in all_doc_ids if any(d.startswith(p) for p in SB_PREFIXES)]
    main_doc_ids = [d for d in all_doc_ids if d not in sb_doc_ids]

    if args.doc_id:
        main_doc_ids = [args.doc_id]

    log_lines = []
    stats = defaultdict(int)

    if sb_doc_ids:
        log_lines.append(
            f"SKIP  {len(sb_doc_ids)} SB County doc(s) excluded (non-standard naming): "
            + ", ".join(sb_doc_ids[:10])
            + (" ..." if len(sb_doc_ids) > 10 else "")
        )

    total = len(main_doc_ids)
    for i, doc_id in enumerate(main_doc_ids):
        json_doc_dir = JSON_DIR / doc_id
        csv_doc_dir  = CSV_DIR  / doc_id

        # Sort JSON files by page number numerically
        json_files = sorted(
            json_doc_dir.glob(f"{doc_id}_page_*.json"),
            key=lambda p: int(re.search(r"_page_(\d+)", p.name).group(1)),
        )
        if not json_files:
            log_lines.append(f"WARN  {doc_id}: no JSON files found, skipping")
            stats["docs_skipped_no_json"] += 1
            continue

        usgs_row = usgs_crosswalk.get(doc_id, {})
        if not usgs_row:
            log_lines.append(f"INFO  {doc_id}: not found in USGS_ID.xlsx crosswalk")
            stats["docs_no_crosswalk"] += 1

        for json_path in json_files:
            m = re.match(rf"{re.escape(doc_id)}_page_(\d+)\.json$", json_path.name)
            if not m:
                log_lines.append(f"WARN  unexpected filename, skipping: {json_path.name}")
                continue
            page_num = m.group(1)

            out_dir = OUTPUT_ROOT / doc_id / f"page_{page_num}"
            if not dry:
                out_dir.mkdir(parents=True, exist_ok=True)

            # JSON
            copy_file(json_path, out_dir / json_path.name, dry)

            # PNG (Jan preferred over Dec)
            png_src = find_png(doc_id, page_num)
            if png_src:
                copy_file(png_src, out_dir / f"{doc_id}_page_{page_num}.png", dry)
            else:
                log_lines.append(f"WARN  {doc_id}/page_{page_num}: PNG not found")
                stats["pages_missing_png"] += 1

            # Table CSVs
            if csv_doc_dir.exists():
                table_csvs = sorted(
                    csv_doc_dir.glob(f"{doc_id}_page_{page_num}_table*.csv")
                )
                for csv_src in table_csvs:
                    copy_file(csv_src, out_dir / csv_src.name, dry)
                if not table_csvs:
                    log_lines.append(f"INFO  {doc_id}/page_{page_num}: no table CSVs found")
                    stats["pages_no_table_csvs"] += 1
            else:
                log_lines.append(f"WARN  {doc_id}: ReductCSVs folder not found")
                stats["docs_no_csv_folder"] += 1

            # Per-page metadata CSV: main metadata rows + USGS doc-level fields
            meta_rows = meta_lookup.get((doc_id, page_num), [])
            if meta_rows:
                page_fields = list(meta_rows[0].keys())
                usgs_extra  = [k for k in usgs_row if k not in page_fields]
                all_fields  = page_fields + usgs_extra
                meta_out    = out_dir / f"{doc_id}_page_{page_num}_metadata.csv"
                if not dry:
                    with open(meta_out, "w", newline="", encoding="utf-8") as mf:
                        writer = csv.DictWriter(
                            mf, fieldnames=all_fields, extrasaction="ignore"
                        )
                        writer.writeheader()
                        for r in meta_rows:
                            writer.writerow({**r, **usgs_row})
            else:
                log_lines.append(
                    f"INFO  {doc_id}/page_{page_num}: no rows in main metadata CSV"
                )
                stats["pages_missing_metadata"] += 1

            stats["pages_processed"] += 1

        stats["docs_processed"] += 1

        if (i + 1) % 50 == 0:
            print(f"  {i + 1}/{total} docs processed...", flush=True)

    # ── Write summary log ─────────────────────────────────────────────────────
    summary_lines = [
        "USGS Data Organization Log",
        "=" * 60,
        f"Total docs processed:            {stats['docs_processed']}",
        f"Total pages processed:           {stats['pages_processed']}",
        f"Pages missing PNG:               {stats['pages_missing_png']}",
        f"Pages missing metadata rows:     {stats['pages_missing_metadata']}",
        f"Pages with no table CSVs:        {stats['pages_no_table_csvs']}",
        f"Docs not in USGS crosswalk:      {stats['docs_no_crosswalk']}",
        f"Docs skipped (no JSON files):    {stats['docs_skipped_no_json']}",
        f"SB County docs excluded:         {len(sb_doc_ids)}",
        "",
        "Detailed notes:",
        *log_lines,
    ]
    log_text = "\n".join(summary_lines)

    log_path = OUTPUT_ROOT / "_organization_log.txt"
    if not dry:
        with open(log_path, "w", encoding="utf-8") as lf:
            lf.write(log_text)
    else:
        print("\n--- Log preview (first 30 lines) ---")
        print("\n".join(summary_lines[:30]))

    print(f"\n{'DRY RUN: ' if dry else ''}Done.")
    print(f"  Docs processed:          {stats['docs_processed']}")
    print(f"  Pages organized:         {stats['pages_processed']}")
    print(f"  Missing PNGs:            {stats['pages_missing_png']}")
    print(f"  Missing metadata rows:   {stats['pages_missing_metadata']}")
    print(f"  Docs not in crosswalk:   {stats['docs_no_crosswalk']}")
    if not dry:
        print(f"\n  Log: {log_path}")


if __name__ == "__main__":
    main()
